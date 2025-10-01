import tkinter as tk
from tkinter import messagebox, ttk
import os
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "materials.xlsx"

def save_data():
    data = [
        entry_name.get(),
        entry_address.get(),
        entry_site.get(),
        entry_mode.get(),
        entry_material.get(),
        entry_size.get(),
        entry_egm.get(),
        entry_msand.get(),
        entry_dust.get(),
        entry_quantity.get()
    ]

    headers = ["Name", "Address", "Site", "Mode", "Type of Material",
               "Size", "EGM", "M. Sand", "Dust", "Quantity (CFT)"]

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(data)
    wb.save(EXCEL_FILE)

    messagebox.showinfo("Saved", f"Data saved to {EXCEL_FILE}")

    # Clear fields
    for entry in [
        entry_name, entry_address, entry_site, entry_mode,
        entry_material, entry_size, entry_egm, entry_msand,
        entry_dust, entry_quantity
    ]:
        entry.delete(0, tk.END)

def view_data():
    if not os.path.exists(EXCEL_FILE):
        messagebox.showwarning("No Data", "No entries saved yet.")
        return

    view_win = tk.Toplevel(root)
    view_win.title("Saved Entries")
    view_win.geometry("1000x500")

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        messagebox.showwarning("No Data", "File is empty.")
        return

    headers = rows[0]
    data_rows = rows[1:]

    # Search bar
    tk.Label(view_win, text="Search:").pack(anchor="w", padx=5, pady=2)
    search_var = tk.StringVar()
    search_entry = tk.Entry(view_win, textvariable=search_var)
    search_entry.pack(fill="x", padx=5, pady=2)

    # Treeview
    tree = ttk.Treeview(view_win, columns=headers, show="headings", selectmode="browse")
    tree.pack(fill=tk.BOTH, expand=True)

    for col in headers:
        tree.heading(col, text=col)
        tree.column(col, width=120)

    def update_table(filter_text=""):
        for row in tree.get_children():
            tree.delete(row)
        for row in data_rows:
            if filter_text.lower() in " ".join([str(cell) for cell in row]).lower():
                tree.insert("", tk.END, values=row)

    update_table()

    def on_key_release(event):
        update_table(search_var.get())

    search_entry.bind("<KeyRelease>", on_key_release)

    # --- Delete Function ---
    def delete_selected():
        selected_item = tree.selection()
        if not selected_item:
            messagebox.showwarning("No Selection", "Please select a row to delete.")
            return

        row_values = tree.item(selected_item)["values"]

        confirm = messagebox.askyesno("Confirm Delete", f"Delete this entry?\n{row_values}")
        if not confirm:
            return

        # Remove from Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        for row in ws.iter_rows(min_row=2):  # skip header
            if [cell.value for cell in row] == list(row_values):
                ws.delete_rows(row[0].row, 1)
                break

        wb.save(EXCEL_FILE)

        # Remove from table
        tree.delete(selected_item)

        messagebox.showinfo("Deleted", "Entry deleted successfully.")

    # Delete button
    tk.Button(view_win, text="Delete Selected", command=delete_selected, bg="red", fg="white").pack(pady=5)

# ---------------- UI ----------------
root = tk.Tk()
root.title("Construction Material Entry")

tk.Label(root, text="Name").grid(row=0, column=0)
entry_name = tk.Entry(root)
entry_name.grid(row=0, column=1)

tk.Label(root, text="Address").grid(row=1, column=0)
entry_address = tk.Entry(root)
entry_address.grid(row=1, column=1)

tk.Label(root, text="Site").grid(row=2, column=0)
entry_site = tk.Entry(root)
entry_site.grid(row=2, column=1)

tk.Label(root, text="Mode").grid(row=3, column=0)
entry_mode = tk.Entry(root)
entry_mode.grid(row=3, column=1)

tk.Label(root, text="Type of Material").grid(row=4, column=0)
entry_material = tk.Entry(root)
entry_material.grid(row=4, column=1)

tk.Label(root, text="Size").grid(row=5, column=0)
entry_size = tk.Entry(root)
entry_size.grid(row=5, column=1)

tk.Label(root, text="EGM").grid(row=6, column=0)
entry_egm = tk.Entry(root)
entry_egm.grid(row=6, column=1)

tk.Label(root, text="M. Sand").grid(row=7, column=0)
entry_msand = tk.Entry(root)
entry_msand.grid(row=7, column=1)

tk.Label(root, text="Dust").grid(row=8, column=0)
entry_dust = tk.Entry(root)
entry_dust.grid(row=8, column=1)

tk.Label(root, text="Quantity (CFT)").grid(row=9, column=0)
entry_quantity = tk.Entry(root)
entry_quantity.grid(row=9, column=1)

# Buttons
tk.Button(root, text="Save Entry", command=save_data).grid(row=10, column=0, columnspan=2, pady=5)
tk.Button(root, text="View Saved Data", command=view_data).grid(row=11, column=0, columnspan=2, pady=5)

root.mainloop()